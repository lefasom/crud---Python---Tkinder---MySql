import tkinter as tk
from tkinter import ttk, messagebox
import csv
import os
from database import GestorBaseDatos
import openpyxl

class Aplicacion:
    def __init__(self):
        self.gestor_bd = GestorBaseDatos()
        self.ventana = tk.Tk()
        self.ventana.geometry("600x400")
        self.ventana.title("CRUD de Alumnos")
        self.ventana.configure(bg="lightblue")

        # Formulario en la mitad izquierda
        self.frame_izquierda = tk.Frame(self.ventana, bg="lightblue")
        self.frame_izquierda.pack(side="left", expand=True, anchor="center")

        self.etiqueta_nombre = tk.Label(self.frame_izquierda, text="Nombre:", font=("Arial", 10), fg="black", bg="lightblue")
        self.etiqueta_nombre.grid(row=0, column=0, pady=(0, 5), sticky="e")

        self.entry_nombre = tk.Entry(self.frame_izquierda)
        self.entry_nombre.grid(row=0, column=1, pady=(0, 10), sticky="w")

        self.etiqueta_edad = tk.Label(self.frame_izquierda, text="Edad:", font=("Arial", 10), fg="black", bg="lightblue")
        self.etiqueta_edad.grid(row=1, column=0, pady=(0, 5), sticky="e")

        self.entry_edad = tk.Entry(self.frame_izquierda)
        self.entry_edad.grid(row=1, column=1, pady=(0, 10), sticky="w")

        self.btn_guardar = tk.Button(self.frame_izquierda, text="Guardar contacto", command=self.guardar_contacto)
        self.btn_guardar.grid(row=2, column=0, columnspan=2, pady=(10, 0))

        self.btn_editar_contacto = tk.Button(self.frame_izquierda, text="Editar contacto", command=self.editar_contacto)
        self.btn_editar_contacto.grid(row=3, column=0, columnspan=2, pady=(10, 0))

        self.btn_exportar_csv = tk.Button(self.frame_izquierda, text="Exportar a CSV", command=self.exportar_a_csv)
        self.btn_exportar_csv.grid(row=4, column=0, columnspan=2, pady=(10, 0))

        # Tabla en la mitad derecha
        self.frame_derecha = tk.Frame(self.ventana, bg="lightblue")
        self.frame_derecha.pack(side="right", expand=True, anchor="center")

        self.tree = ttk.Treeview(self.frame_derecha, columns=("Id", "Nombre", "Edad"), show="headings")
        self.tree.heading("Id", text="Id", anchor="w")
        self.tree.heading("Nombre", text="Nombre", anchor="w")
        self.tree.heading("Edad", text="Edad", anchor="w")
        self.tree.grid(row=0, column=0, rowspan=3, padx=10, sticky="nsew")

        self.btn_mostrar_contactos = tk.Button(self.frame_derecha, text="Mostrar contactos", command=self.mostrar_contactos)
        self.btn_mostrar_contactos.grid(row=4, column=0, pady=(10, 0))

        self.btn_borrar_contacto = tk.Button(self.frame_derecha, text="Borrar contacto", command=self.borrar_contacto)
        self.btn_borrar_contacto.grid(row=3, column=0, pady=(10, 0))

        self.btn_exportar_xlsx = tk.Button(self.frame_izquierda, text="Exportar a XLSX", command=self.exportar_a_xlsx)
        self.btn_exportar_xlsx.grid(row=5, column=0, columnspan=2, pady=(10, 0))
        
        # Configuraciones adicionales para el Treeview
        self.tree.column("Id", width=20, anchor="center")
        self.tree.column("Nombre", width=120, anchor="center")
        self.tree.column("Edad", width=50, anchor="center")
        self.tree.bind("<ButtonRelease-1>", self.seleccionar_contacto)

        # Cerrar la ventana al cerrar la aplicación
        self.ventana.protocol("WM_DELETE_WINDOW", self.ventana.destroy)

    def mostrar_contactos(self):
        """Mostrar los contactos en la tabla."""
        # Limpiar la tabla antes de volver a cargar los datos
        for item in self.tree.get_children():
            self.tree.delete(item)

        contactos = self.gestor_bd.obtener_contactos()

        for contacto in contactos:
            id, nombre, edad = contacto
            valores = (id, nombre, edad)
            self.tree.insert("", "end", values=valores)

    def guardar_contacto(self):
        """Guardar un nuevo contacto desde la interfaz gráfica."""
        nombre = self.entry_nombre.get()
        edad = self.entry_edad.get()

        # Validar que la edad sea un número entero
        try:
            edad = int(edad)
        except ValueError:
            messagebox.showerror("Error", "La edad debe ser un número entero.")
            return

        self.gestor_bd.guardar_contacto(nombre, edad)
        # messagebox.showinfo("Éxito", "Contacto guardado exitosamente.")
        self.mostrar_contactos()  # Actualizar la tabla después de guardar

    def borrar_contacto(self):
        """Borrar el contacto seleccionado."""
        item = self.tree.selection()
        if item:
            id = self.tree.item(item, "values")[0]
            self.gestor_bd.eliminar_contacto(id)
            self.mostrar_contactos()  # Actualizar la tabla después de borrar

    def seleccionar_contacto(self, event):
        """Obtener el contacto seleccionado en la tabla."""
        item = self.tree.selection()
        if item:
            id = self.tree.item(item, "values")[0]
            nombre = self.tree.item(item, "values")[1]
            edad = self.tree.item(item, "values")[2]
            mensaje = f"Nombre: {nombre}, Edad: {edad}"
            # messagebox.showinfo("Contacto seleccionado", mensaje)

    def editar_contacto(self):
        """Abrir una nueva ventana para editar el contacto seleccionado."""
        item = self.tree.selection()
        if item:
            id = self.tree.item(item, "values")[0]
            nombre = self.tree.item(item, "values")[1]
            edad = self.tree.item(item, "values")[2]

            ventana_edicion = tk.Toplevel(self.ventana, bg="lightblue")
            ventana_edicion.title("Editar contacto")
            ventana_edicion.geometry("300x150")

            # Crear widgets para la ventana de edición
            label_nombre = tk.Label(ventana_edicion, text="Nuevo nombre:", bg="lightblue")
            label_nombre.pack()

            entry_nuevo_nombre = tk.Entry(ventana_edicion)
            entry_nuevo_nombre.insert(0, nombre)
            entry_nuevo_nombre.pack()

            label_edad = tk.Label(ventana_edicion, text="Nueva edad:", bg="lightblue")
            label_edad.pack()

            entry_nueva_edad = tk.Entry(ventana_edicion)
            entry_nueva_edad.insert(0, edad)
            entry_nueva_edad.pack()

            btn_guardar_edicion = tk.Button(ventana_edicion, text="Guardar cambios", command=lambda: self.guardar_edicion(id, entry_nuevo_nombre.get(), entry_nueva_edad.get(), ventana_edicion))
            btn_guardar_edicion.pack()

    def guardar_edicion(self, id, nuevo_nombre, nueva_edad, ventana_edicion):
        """Guardar los cambios después de editar un contacto."""
        # Validar que la nueva edad sea un número entero
        try:
            nueva_edad = int(nueva_edad)
        except ValueError:
            messagebox.showerror("Error", "La nueva edad debe ser un número entero.")
            return

        self.gestor_bd.actualizar_contacto(id, nuevo_nombre, nueva_edad)
        messagebox.showinfo("Éxito", "Cambios guardados exitosamente.")
        ventana_edicion.destroy()
        self.mostrar_contactos()  # Actualizar la tabla después de editar

    def exportar_a_csv(self):
        """Exportar los datos a un archivo CSV."""
        contactos = self.gestor_bd.obtener_contactos()
        encabezados = ["Id", "Nombre", "Edad"]

        # Obtener el directorio del escritorio
        escritorio = os.path.join(os.path.expanduser('~'), 'Desktop')

        with open(os.path.join(escritorio, "contactos.csv"), mode="w", newline="", encoding="utf-8") as archivo_csv:
            escritor_csv = csv.writer(archivo_csv)
            
            # Escribir encabezados
            escritor_csv.writerow(encabezados)

            # Escribir datos
            escritor_csv.writerows(contactos)

        messagebox.showinfo("Éxito", "Datos exportados a contactos.csv en el escritorio")

    def exportar_a_xlsx(self):
        """Exportar los datos a un archivo xlsx."""
        contactos = self.gestor_bd.obtener_contactos()
        encabezados = ["Id", "Nombre", "Edad"]

        # Obtener el directorio del escritorio
        escritorio = os.path.join(os.path.expanduser('~'), 'Desktop')

        # Crear un nuevo libro y seleccionar la hoja activa
        libro = openpyxl.Workbook()
        hoja = libro.active

        # Escribir encabezados en la primera fila
        for col_num, encabezado in enumerate(encabezados, 1):
            hoja.cell(row=1, column=col_num, value=encabezado)

        # Escribir datos en filas sucesivas
        for fila_num, contacto in enumerate(contactos, 2):
            for col_num, valor in enumerate(contacto, 1):
                hoja.cell(row=fila_num, column=col_num, value=valor)

        # Guardar el libro en el escritorio
        libro.save(os.path.join(escritorio, "contactos.xlsx"))

        messagebox.showinfo("Éxito", "Datos exportados a contactos.xlsx en el escritorio")
    def ejecutar(self):
        """Iniciar la aplicación."""
        self.ventana.mainloop()

if __name__ == "__main__":
    app = Aplicacion()
    app.ejecutar()
